#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel Exporter - Xuất dữ liệu sang Excel theo mẫu
"""

import os
import logging
from pathlib import Path
try:
    from .io_safety import assert_not_pdf_target, IOErrorSafety
except ImportError:
    from io_safety import assert_not_pdf_target, IOErrorSafety
from typing import List, Dict, Any, Optional
try:
    from .models import DocumentMetadata
except Exception:
    # Support running module as a script (no parent package)
    try:
        from models import DocumentMetadata
    except Exception:
        # Last-resort: attempt import via src package name
        from src.models import DocumentMetadata

# Configure logging
logger = logging.getLogger(__name__)


class ExcelExportError(Exception):
    """Base exception for Excel export errors"""
    pass


class DependencyError(ExcelExportError):
    """Raised when required dependencies are missing"""
    pass


class FileAccessError(ExcelExportError):
    """Raised when file cannot be accessed or written"""
    pass


class DataValidationError(ExcelExportError):
    """Raised when input data is invalid"""
    pass


class ExcelExporter:
    """Xuất metadata sang Excel theo mẫu chuẩn

    This class lazily imports pandas and openpyxl in __init__ and exposes a
    single `export()` method that is safe (checks permissions, handles file
    locks, and formats the workbook). Known exceptions are raised as
    subclasses of ExcelExportError.
    """

    def __init__(self) -> None:
        # Lazy imports and style setup
        try:
            import pandas as pd  # type: ignore
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # type: ignore
        except Exception as e:  # pragma: no cover - environment dependent
            logger.exception("Missing Excel runtime dependencies")
            raise DependencyError(
                "Missing required dependencies: pandas and openpyxl.\n"
                "Install with: pip install pandas openpyxl"
            ) from e

        self.pd = pd
        self.Font = Font
        self.PatternFill = PatternFill
        self.Border = Border
        self.Side = Side
        self.Alignment = Alignment

        # Styles (header per ideal.txt: Times New Roman size 13 bold)
        self.header_font = Font(name='Times New Roman', size=13, bold=True)
        # Body uses Times New Roman size 13 (regular)
        self.body_font = Font(name='Times New Roman', size=13)
        self.header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        self.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _validate_metadata(self, metadata_list: List[DocumentMetadata]) -> None:
        if not metadata_list:
            raise DataValidationError("Metadata list is empty")
        # Accept either DocumentMetadata instances or plain dicts for backward
        # compatibility with other parts of the code that produce dicts.
        for i, m in enumerate(metadata_list):
            if not (isinstance(m, DocumentMetadata) or isinstance(m, dict)):
                raise DataValidationError(f"Item {i} must be DocumentMetadata or dict, got: {type(m)}")

    def _check_file_access(self, output_path: str) -> None:
        path = Path(output_path)
        # Prevent accidental writes directly to PDF files
        try:
            assert_not_pdf_target(str(path))
        except IOErrorSafety as e:
            raise FileAccessError(str(e)) from e
        if not path.parent.exists():
            try:
                path.parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                raise FileAccessError(f"Cannot create directory {path.parent}: {e}") from e

        if not os.access(path.parent, os.W_OK):
            raise FileAccessError(f"No write permission for directory: {path.parent}")

        # If file exists, attempt to open for append to check locks
        if path.exists():
            try:
                with open(path, 'a', encoding='utf-8'):
                    pass
            except Exception as e:
                raise FileAccessError(f"Cannot write to {path}. File may be open: {e}") from e

    def _records_from_metadata(self, metadata_list: List[DocumentMetadata]) -> List[Dict[str, Any]]:
        records: List[Dict[str, Any]] = []
        for idx, meta in enumerate(metadata_list, start=1):
            def get(k1: str, *alternatives: str):
                if not isinstance(meta, dict):
                    val = getattr(meta, k1, None)
                    if val:
                        return val
                    for alt in alternatives:
                        val = getattr(meta, alt, None)
                        if val:
                            return val
                    return ''
                if k1 in meta and meta.get(k1) is not None:
                    return meta.get(k1)
                for alt in alternatives:
                    if alt in meta and meta.get(alt) is not None:
                        return meta.get(alt)
                return ''

            try:
                rec: Dict[str, Any] = {
                    'STT': idx,
                    'Tiêu đề hồ sơ': get('tieu_de', 'tieu_de_ho_so'),
                    'Số hồ sơ': get('so_ho_so'),
                    'Thời gian bắt đầu': get('thoi_gian_bat_dau', 'tu_ngay'),
                    'Thời gian kết thúc': get('thoi_gian_ket_thuc', 'den_ngay'),
                    'Số lượng tờ': get('so_to', 'so_luong_to'),
                    'Thời hạn bảo quản': get('thoi_han_bao_quan', 'thoi_han'),
                    'Tên phông': get('ten_phong'),
                    'Mã phông': get('ma_phong'),
                    'Hộp Số': get('hop_so', 'hop'),
                    'Xem': get('xem_file'),
                    'Nhiệm kỳ': get('nhiem_ky'),
                    'Số lượng trang hồ sơ': get('so_trang', 'so_trang_hs'),
                    'Tổng số văn bản trong hồ sơ': get('tong_so_van_ban', 'tong_so'),
                    # Keep organization value as-is (don't require ALL uppercase)
                    'Cơ quan ban hành': get('co_quan_ban_hanh', 'co_quan'),
                    'Số văn bản': get('so_van_ban', 'so_van'),
                    'Ký hiệu': get('ky_hieu_van_ban', 'ky_hieu'),
                    'Ngày ký': get('ngay_ky', 'ngay'),
                    'Thể loại văn bản': get('the_loai_van_ban', 'the_loai'),
                    'Trích yếu nội dung văn bản': get('trich_yeu_noi_dung', 'trich_yeu'),
                    'Người Ký': get('nguoi_ky'),
                    'Loại văn bản': get('loai_ban', 'loai_van_ban'),
                    'Trang Số': get('trang_so', 'trang_bat_dau'),
                    'Số trang': get('so_trang_van_ban', 'so_trang'),
                    'Địa chỉ tài liệu gốc': get('dia_chi_tai_lieu', 'duong_dan_file', 'duong_dan'),
                    'Người nhập tin': get('nguoi_nhap_tin', 'nguoi_nhap')
                }
            except Exception:
                logger.exception("Skipping record due to processing error")
                continue
            # Previously we removed many fields when 'Thể loại văn bản' was empty.
            # That behavior caused valid extracted fields to be discarded when
            # the extractor couldn't confidently determine the document type.
            # Keep all extracted fields as-is to preserve information; callers
            # can filter or validate downstream if needed.
            records.append(rec)
        return records

    def _format_worksheet(self, worksheet) -> None:
        try:
            # Header
            for cell in worksheet[1]:
                try:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.border = self.border
                    cell.alignment = self.center_aligned
                except Exception:
                    continue

            # Body - simple formatting, no bolding
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    try:
                        cell.font = self.body_font
                        cell.border = self.border
                    except Exception:
                        continue

            # Column widths
            for col in worksheet.columns:
                max_len = 0
                col = list(col)
                for cell in col:
                    try:
                        val = cell.value
                        length = len(str(val)) if val is not None else 0
                        if length > max_len:
                            max_len = length
                    except Exception:
                        continue
                width = min(max_len + 2, 50)
                try:
                    worksheet.column_dimensions[col[0].column_letter].width = width
                except Exception:
                    continue

            worksheet.freeze_panes = 'A2'
            # Try to apply number/date formats for common columns if present
            try:
                # Map column headers to number formats
                formats = {
                    'Ngày ký': 'DD/MM/YYYY',
                    'Số trang': '0',
                    'Trang Số': '0',
                    'Số trang': '0',
                    'Số lượng tờ': '0'
                }
                header = [c.value for c in worksheet[1]]
                for col_idx, col_name in enumerate(header, start=1):
                    fmt = formats.get(col_name)
                    if not fmt:
                        continue
                    for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=worksheet.max_row):
                        cell = row[0]
                        try:
                            cell.number_format = fmt
                        except Exception:
                            continue
            except Exception:
                # don't fail the whole formatting pass on number-formatting issues
                logger.exception("Applying number/date formats failed")
        except Exception:
            logger.exception("Worksheet formatting failed but continuing")

    def export(self, metadata_list: List[DocumentMetadata], output_path: str, 
              sheet_name: str = 'Biên mục', batch_size: int = 1000) -> str:
        """Export metadata_list to an Excel file.

        Args:
            metadata_list: List of metadata records
            output_path: Path to output Excel file
            sheet_name: Name of worksheet
            batch_size: Number of records to process in each batch for memory efficiency

        Raises ExcelExportError subclasses on known failures.
        """
        # Validate input & file accessibility
        self._validate_metadata(metadata_list)
        self._check_file_access(output_path)
        
        # Process in batches for large datasets
        total_records = len(metadata_list)
        records = []
        
        if total_records > batch_size:
            for i in range(0, total_records, batch_size):
                batch = metadata_list[i:i + batch_size]
                batch_records = self._records_from_metadata(batch)
                records.extend(batch_records)
                logger.info(f"Processed batch {i//batch_size + 1}/{(total_records + batch_size - 1)//batch_size}")
        else:
            records = self._records_from_metadata(metadata_list)
            
        if not records:
            raise DataValidationError('No valid records to export')

        try:
            # Define exact column order per user specification (Vietnamese names and order)
            columns = [
                'STT',
                'Tiêu đề hồ sơ',
                'Số hồ sơ',
                'Thời gian bắt đầu',
                'Thời gian kết thúc',
                'Số lượng tờ',
                'Thời hạn bảo quản',
                'Tên phông',
                'Mã phông',
                'Hộp Số',
                'Số lượng trang hồ sơ',
                'Tổng số văn bản trong hồ sơ',
                # Đưa đường dẫn, hyperlink và nhiệm kỳ về gần thông tin cơ quan ban hành
                'Xem',
                'Nhiệm kỳ',
                'Cơ quan ban hành',
                'Số văn bản',
                'Ký hiệu',
                'Ngày ký',
                'Thể loại văn bản',
                'Trích yếu nội dung văn bản',
                'Người Ký',
                'Loại văn bản',
                'Trang Số',
                'Số trang',
                'Địa chỉ tài liệu gốc',
                'Người nhập tin'
            ]
            
            # Create DataFrame
            df = self.pd.DataFrame(records)
            
            # --- Implement Page Numbering Logic (User Request) ---
            # 1. Ensure 'Số trang' is numeric (fill invalid/empty with 0)
            if 'Số trang' in df.columns:
                df['Số trang'] = self.pd.to_numeric(df['Số trang'], errors='coerce').fillna(0).astype(int)
            else:
                 df['Số trang'] = 0
            
            # 2. Sort by 'Số hồ sơ' and 'Số văn bản' (or 'STT' if implicit)
            # 'Số văn bản' might be text (e.g. "01/NQ"), we want to sort roughly by number.
            # But sorting by just 'Số hồ sơ' is most critical for grouping.
            # We assume the incoming list order is roughly correct or we sort by 'Số văn bản'.
            # To sort by 'Số văn bản' correctly (numeric), we might need auxiliary column.
            # For now, simplistic sort:
            try:
                if 'Số hồ sơ' in df.columns:
                     # Create temporary sort column for Doc Number if possible?
                     # Just sort standard.
                     df = df.sort_values(by=['Số hồ sơ', 'STT']) # Maintain relative order from input list if no better key?
                     # Actually User said "vòng lặp thực hiện nếu cùng số hồ sơ".
                     # Assuming input list `metadata_list` is already sorted by file/folder structure (which usually implies date/number).
                     # So separating by 'Số hồ sơ' is enough.
                     pass 
            except Exception:
                pass
            
            # 3. Group by 'Số hồ sơ' and calculate Start Page ('Trang Số')
            # Start Page = 1 + Shifted Cumulative Sum of Page Count
            if 'Số hồ sơ' in df.columns and 'Số trang' in df.columns:
                # Groupby Dossier
                # Function to calc pages per group
                def calc_start_pages(group):
                    # Shift: The start page of row i is (Sum of pages 0..i-1) + 1
                    # cumsum includes current row. so shift it.
                    # e.g. Pages: [2, 3, 5]
                    # cumsum: [2, 5, 10]
                    # shift(1): [NaN, 2, 5]
                    # fillna(0): [0, 2, 5]
                    # +1: [1, 3, 6] -> Correct.
                    return group['Số trang'].cumsum().shift(1).fillna(0).astype(int) + 1
                
                # Apply transformation
                try:
                    df['Trang Số'] = df.groupby('Số hồ sơ', group_keys=False).apply(calc_start_pages)
                except Exception as e:
                    logger.warning(f"Failed to calculate page numbering: {e}")
                    # Fallback or leave as is?
                    # df['Trang Số'] = 1
                    pass

            
            # Ensure column order (only keep columns that exist after reindex)
            # Reindex to include missing columns with empty values
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            # Force exact column order and fill NaN with empty strings
            df = df.reindex(columns=columns)
            df = df.fillna("")

            with self.pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                # Apply formatting
                self._format_worksheet(worksheet)

            logger.info("Exported %d records to %s", len(records), output_path)
            return output_path
        except Exception as e:
            logger.exception("Failed to export Excel")
            raise ExcelExportError(f"Failed to export Excel file: {e}") from e

    def export_metadata_dicts(self, metadata_list: List[Dict[str, str]], output_path: str,
                              sheet_name: str = 'Metadata') -> str:
        """Export a simple list of metadata dicts to Excel.

        Each dict should contain the keys: issuer, doc_number, symbol, signed_date,
        doc_type, summary, signer. Missing keys will be exported as empty strings.

        This produces one row per PDF and auto-adjusts column widths. Unicode is
        preserved (openpyxl/pandas handle UTF-8 for Excel files).
        """
        # Basic validation
        if not isinstance(metadata_list, list):
            raise DataValidationError('metadata_list must be a list of dicts')

        # Normalize list of dicts
        rows = []
        for i, item in enumerate(metadata_list):
            if not isinstance(item, dict):
                raise DataValidationError(f'Item {i} is not a dict')
            # ensure all values are strings (or empty)
            row = {
                'issuer': str(item.get('issuer', '') or ''),
                'doc_number': str(item.get('doc_number', '') or ''),
                'symbol': str(item.get('symbol', '') or ''),
                'signed_date': str(item.get('signed_date', '') or ''),
                'doc_type': str(item.get('doc_type', '') or ''),
                'summary': str(item.get('summary', '') or ''),
                'signer': str(item.get('signer', '') or ''),
            }
            rows.append(row)

        # Convert to DataFrame and write to Excel
        try:
            df = self.pd.DataFrame(rows)
            # Ensure column order
            columns = ['issuer', 'doc_number', 'symbol', 'signed_date', 'doc_type', 'summary', 'signer']
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            df = df.reindex(columns=columns)
            df = df.fillna("")

            self._check_file_access(output_path)

            with self.pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                # Apply basic formatting and adjust column widths
                try:
                    # reuse the existing formatting helper for consistent look
                    self._format_worksheet(worksheet)
                except Exception:
                    logger.exception('Failed to format metadata worksheet')

            logger.info('Exported metadata dicts to %s', output_path)
            return output_path
        except Exception as e:
            logger.exception('Failed to export metadata dicts')
            raise ExcelExportError(f'Failed to export metadata dicts: {e}') from e
